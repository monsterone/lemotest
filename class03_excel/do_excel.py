# -*- coding: utf-8 -*-
# @Time     : 2020/4/4 16:07
# @Author   : Monster
# @Email    : 945534456@qq.com
# @File     : do_excel.py

from openpyxl import load_workbook

class DoExcel():

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]  #错误点二：不要写成字符串
        test_data=[]
        for i in range(2,sheet.max_row+1):
            sub_data = {}
            sub_data['method'] = sheet.cell(i,1).value
            sub_data['url'] = sheet.cell(i,2).value
            sub_data['data'] = sheet.cell(i,3).value
            sub_data['expected'] = sheet.cell(i,4).value
            test_data.append(sub_data)
        return test_data

if __name__ == '__main__':
    res = DoExcel(r"C:\Users\asus\Desktop\test_header.xlsx","python").get_data()
    print(res)