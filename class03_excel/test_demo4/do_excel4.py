# -*- coding: utf-8 -*-
# @Time     : 2020/4/4 16:07
# @Author   : Monster
# @Email    : 945534456@qq.com
# @File     : do_excel.py

from openpyxl import load_workbook

#方法三：仅供参考，日后拿来优化
class DoExcel():

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def get_header(self):
        '''获取第一行的标题行'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = [] #存储我们的标题行
        for j in range(1,sheet.max_column+1):
            header.append(sheet.cell(1,j).value)
        return header



    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        header = self.get_header()  #拿到header
        test_data=[]
        for i in range(2,sheet.max_row+1): #（1,5）1 2 3 4
            sub_data = {}
            for j in range(1,sheet.max_column+1): #1 2 3 4 5...
                sub_data[header[j-1]] = sheet.cell(i,j).value  #head是列表 ，所以j-1
            test_data.append(sub_data)

        return test_data #返回获取到的数据

if __name__ == '__main__':
    res = DoExcel(r"C:\Users\asus\Desktop\test_header.xlsx","python").get_data()
    print(res)