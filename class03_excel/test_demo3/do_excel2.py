# -*- coding: utf-8 -*-
# @Time     : 2020/4/4 16:07
# @Author   : Monster
# @Email    : 945534456@qq.com
# @File     : do_excel.py

from openpyxl import load_workbook

class DoExcel():

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
        self.sheet_obj=load_workbook(self.file_name)[self.sheet_name]  #获取一个表单对象
        self.max_row=self.sheet_obj.max_row #获取表单对象最大行数

    def get_data(self,i,j):
       '''根据传入的坐标来取值'''
       return self.sheet_obj.cell(i,j).value

if __name__ == '__main__':
    res = DoExcel(r"C:\Users\asus\Desktop\test.xlsx","python").get_data(1,1)
    print(res)