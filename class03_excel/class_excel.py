# -*- coding: utf-8 -*-
# @Time     : 2020/4/2 22:32
# @Author   : Monster
# @Email    : 945534456@qq.com
# @File     : class_excel.py

#地址 测试数据 断言期望结果 除了这几个不同 其他的都是高度相识 80%
#参数化 url data expected ---参考：test_demo

#----------

#存到excel里面 python去操作Excel
#1:只支持这种后缀：.xlsx  ---> openpyxl 只支持这种格式
#2:老老实实创建

from openpyxl import load_workbook

#1:打开Excel
# wb = load_workbook(r"C:\Users\asus\Desktop\test.xlsx")

#2:定位表单
# sheet = wb["python"]  #传表单名 返回一个表单对象

#3:定位单元格  行列值定位
# res = sheet.cell(1,1).value


# print("最大行:{}".format(sheet.max_row))    #求表单最大行
# print("最大列:{}".format(sheet.max_column)) #求表单的最大列
# print("拿到的结果是:",res)

# ---------------
#数据从Excel里面拿出来是什么类型
#数字还是数字，其他都是字符串

#eveal() 把数据转化成 原本数据类型

#错误演示一：只定位到单元格 并未取值
# res = sheet.cell(1,1)

#怎么拿到第一行的所有数据
wb = load_workbook(r"C:\Users\asus\Desktop\test.xlsx")
sheet = wb["python"]
test_data=[]
for i in range(1,sheet.max_row+1):
    sub_data = {}
    # method = sheet.cell(i,1).value
    sub_data['method'] = sheet.cell(i,1).value
    # url = sheet.cell(i,2).value
    sub_data['url'] = sheet.cell(i,2).value
    # data = sheet.cell(i,3).value
    sub_data['data'] = sheet.cell(i,3).value
    # expected = sheet.cell(i,4).value
    sub_data['expected'] = sheet.cell(i,4).value
    test_data.append(sub_data)
print(test_data)