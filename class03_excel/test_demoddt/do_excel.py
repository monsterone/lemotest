# -*- coding: utf-8 -*-
# @Time     : 2020/4/4 16:07
# @Author   : Monster
# @Email    : 945534456@qq.com
# @File     : do_excel.py

from openpyxl import load_workbook

from class03_excel.do_config.read_config import ReadConfig
class DoExcel():

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def get_data(self,mode='all'):
    # def get_data(self):
        '''mode:控制是否执行所有的用例 默认为all 为all就执行所以用例
        如果不等于all的话 就进入分值判断
        mode的值 只输入 all 列表 这两种类型的参数
        :return:
        '''

        # #从配置文件读取mode值
        # mode=ReadConfig().read_config('case.config','MODE','mode')
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]  #错误点二：不要写成字符串
        test_data=[]
        for i in range(2,sheet.max_row+1):
            sub_data = {}
            sub_data['case_id'] = sheet.cell(i, 1).value
            sub_data['module'] = sheet.cell(i, 2).value
            sub_data['title'] = sheet.cell(i, 3).value
            sub_data['method'] = sheet.cell(i,4).value
            sub_data['url'] = sheet.cell(i,5).value
            sub_data['data'] = sheet.cell(i,6).value
            sub_data['expected'] = sheet.cell(i,7).value
            test_data.append(sub_data)

        #根据mode值去进行判断
        if mode == 'all':
            final_data=test_data
        else:#[1,2,3]
            final_data=[]
            for item in test_data: #对test_data 所以的测试数据 进行遍历
                if item['case_id'] in eval(mode):
                    final_data.append(item)

        return final_data  #返回获取到的数据

if __name__ == '__main__':
    res = DoExcel(r"C:\Users\asus\Desktop\test_header.xlsx","python").get_data()
    print(res)